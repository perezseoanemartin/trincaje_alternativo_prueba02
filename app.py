import streamlit as st
import formulas
import os
import numpy as np
import pandas as pd
import openpyxl  # Necesario para leer las fórmulas

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Trincaje Pro Final", page_icon="⚓", layout="wide")
st.title("⚓ Calculadora de Trincaje (Con Inspector)")

# CONSTANTES
NOMBRE_HOJA = "CALCULO"  
ARCHIVO_EXCEL = "trincaje_alternativo_prueba02.xlsx"

# DEFINICIÓN DE MATERIALES
MAPA_MATERIALES = [
    {"nombre": "Grillete/Tensor",    "fila": 64, "factor": 0.5}, 
    {"nombre": "Cabo Fibra",         "fila": 65, "factor": 0.33}, 
    {"nombre": "Cable 1 uso",        "fila": 66, "factor": 0.8}, 
    {"nombre": "Cable Reutiliz.",    "fila": 67, "factor": 0.3}, 
    {"nombre": "Fleje acero",        "fila": 68, "factor": 0.7}, 
    {"nombre": "Cincha",             "fila": 69, "factor": 0.5}, 
    {"nombre": "Bulldog Grip",       "fila": 70, "factor": 0.7}, 
    {"nombre": "Madera",             "fila": 71, "factor": 0.3}
]

# --- 1. CARGA DEL MOTOR ---
@st.cache_resource
def cargar_motor():
    if not os.path.exists(ARCHIVO_EXCEL):
        return None
    # Cargamos motor de cálculo
    xl_model = formulas.ExcelModel().loads(ARCHIVO_EXCEL).finish()
    return xl_model

try:
    modelo = cargar_motor()
except Exception as e:
    st.error(f"Error cargando el motor: {e}")
    st.stop()

if not modelo:
    st.error(f"⚠️ Error: No encuentro '{ARCHIVO_EXCEL}'.")
    st.stop()

# --- 2. INTERFAZ: DATOS GENERALES ---
with st.expander("🚢 1. Datos del Buque y Carga", expanded=True):
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown("**Buque**")
        c_eslora = st.number_input("Eslora (C64)", value=0.0)
        c_velocidad = st.number_input("Velocidad (C65)", value=0.0)
        c_corr = st.number_input("Corrección Tabla 4 (C67)", value=0.0)
    with col2:
        st.markdown("**Aceleraciones**")
        c_ax = st.number_input("ax (C69)", value=0.0)
        c_ay = st.number_input("ay (C70)", value=0.0)
        c_az = st.number_input("az (C71)", value=0.0)
    with col3:
        st.markdown("**Carga**")
        e_eslora = st.number_input("Eslora Carga (E64)", value=0.0)
        e_manga = st.number_input("Manga Carga (E65)", value=0.0)
        e_altura = st.number_input("Altura Carga (E66)", value=0.0)
        e_masa = st.number_input("Masa Carga (E67)", value=0.0)
        e_friccion = st.number_input("Coef. Fricción (E68)", value=0.0)
    
    st.markdown("**Brazos de Estabilidad**")
    c_b1, c_b2, c_b3 = st.columns(3)
    e_brazo_v = c_b1.number_input("Brazo Vuelco (E69)", value=0.0)
    e_brazo_br = c_b2.number_input("Brazo Estab. Br (E70)", value=0.0)
    e_brazo_er = c_b3.number_input("Brazo Estab. Er (E71)", value=0.0)

# --- 3. RESISTENCIA DE MATERIALES ---
with st.expander("🛠️ 2. Resistencia de Materiales", expanded=True):
    st.info("Introduce los valores G y H.")
    
    inputs_para_excel_mat = {}
    valores_k_finales = {"-": 0.0}
    
    cols_mat = st.columns(4)
    
    for i, item in enumerate(MAPA_MATERIALES):
        nombre = item["nombre"]
        fila = item["fila"]
        factor = item["factor"]
        
        with cols_mat[i % 4]:
            st.markdown(f"**{nombre}**")
            val_g = st.number_input(f"G{fila}", key=f"G{fila}", value=0.0, step=0.1)
            val_h = st.selectbox(f"H{fila}", ["-", "Tm", "KN"], key=f"H{fila}", index=0)
            
            k_calc = 0.0
            if val_h == "Tm":
                k_calc = val_g * factor * 9.8
            elif val_h == "KN":
                k_calc = val_g * factor
            
            if k_calc > 0:
                st.markdown(f":green[**= {k_calc:.2f} KN**]")
            else:
                st.caption("= 0.00 KN")
                
            inputs_para_excel_mat[f"G{fila}"] = val_g
            inputs_para_excel_mat[f"H{fila}"] = val_h
            valores_k_finales[nombre] = k_calc

# --- 4. CONFIGURACIÓN DE TRINCAS ---
st.subheader("⛓️ 3. Configuración de Trincas")
opciones_desplegable = ["-"] + [m["nombre"] for m in MAPA_MATERIALES]

tab_stbd, tab_port = st.tabs(["Estribor (Starboard)", "Babor (Portside)"])

datos_trincas_estribor = []
datos_trincas_babor = []

# ESTRIBOR
with tab_stbd:
    cols = st.columns([3, 1.5, 1.5, 1.5, 1.5])
    cols[0].write("Material"); cols[1].write("Brazo C (E)"); cols[2].write("Alfa (F)"); cols[3].write("Beta (G)"); cols[4].write("Dir (H)")
    
    for i in range(6):
        fila = 86 + i
        c1, c2, c3, c4, c5 = st.columns([3, 1.5, 1.5, 1.5, 1.5])
        
        sel = c1.selectbox(f"Trinca #{i+1}", opciones_desplegable, key=f"st_Mat_{fila}", label_visibility="collapsed")
        val_k_real = valores_k_finales.get(sel, 0.0)
        
        brazo = c2.number_input("Brazo", 0.0, key=f"st_Brazo_{fila}", label_visibility="collapsed")
        alfa = c3.number_input("Alfa", 0.0, key=f"st_Alfa_{fila}", label_visibility="collapsed")
        beta = c4.number_input("Beta", 0.0, key=f"st_Beta_{fila}", label_visibility="collapsed")
        dire = c5.selectbox("Dir", ["-", "Pr", "Pp"], key=f"st_Dir_{fila}", label_visibility="collapsed")
        
        datos_trincas_estribor.append({"fila": fila, "B": val_k_real, "Brazo": brazo, "F": alfa, "G": beta, "H": dire})

# BABOR
with tab_port:
    cols = st.columns([3, 1.5, 1.5, 1.5, 1.5])
    cols[0].write("Material"); cols[1].write("Brazo C (C)"); cols[2].write("Alfa (F)"); cols[3].write("Beta (G)"); cols[4].write("Dir (H)")
    
    for i in range(6):
        fila = 93 + i
        c1, c2, c3, c4, c5 = st.columns([3, 1.5, 1.5, 1.5, 1.5])
        
        sel = c1.selectbox(f"Trinca #{i+1}", opciones_desplegable, key=f"pt_Mat_{fila}", label_visibility="collapsed")
        val_k_real = valores_k_finales.get(sel, 0.0)
        
        brazo = c2.number_input("Brazo", 0.0, key=f"pt_Brazo_{fila}", label_visibility="collapsed")
        alfa = c3.number_input("Alfa", 0.0, key=f"pt_Alfa_{fila}", label_visibility="collapsed")
        beta = c4.number_input("Beta", 0.0, key=f"pt_Beta_{fila}", label_visibility="collapsed")
        dire = c5.selectbox("Dir", ["-", "Pr", "Pp"], key=f"pt_Dir_{fila}", label_visibility="collapsed")
        
        datos_trincas_babor.append({"fila": fila, "B": val_k_real, "Brazo": brazo, "F": alfa, "G": beta, "H": dire})

# --- 5. CÁLCULO ---
if st.button("🚀 Calcular Seguridad", type="primary"):
    
    inputs_dict = {}
    def add(celda, valor):
        key = f"'[{ARCHIVO_EXCEL}]{NOMBRE_HOJA}'!{celda}"
        inputs_dict[key] = valor

    try:
        # A) General
        add("C64", c_eslora); add("C65", c_velocidad); add("C67", c_corr)
        add("C69", c_ax); add("C70", c_ay); add("C71", c_az)
        add("E64", e_eslora); add("E65", e_manga); add("E66", e_altura)
        add("E67", e_masa); add("E68", e_friccion); add("E69", e_brazo_v)
        add("E70", e_brazo_br); add("E71", e_brazo_er)

        # B) Materiales
        for c, v in inputs_para_excel_mat.items(): add(c, v)

        # C) Estribor
        for t in datos_trincas_estribor:
            f = t['fila']
            add(f"B{f}", t['B']); add(f"E{f}", t['Brazo']); add(f"F{f}", t['F']); add(f"G{f}", t['G']); add(f"H{f}", t['H'])

        # D) Babor
        for t in datos_trincas_babor:
            f = t['fila']
            add(f"B{f}", t['B']); add(f"C{f}", t['Brazo']); add(f"F{f}", t['F']); add(f"G{f}", t['G']); add(f"H{f}", t['H'])

        # --- EJECUCIÓN ---
        solution = modelo.calculate(inputs=inputs_dict)

        def get(celda):
            key = f"'[{ARCHIVO_EXCEL}]{NOMBRE_HOJA}'!{celda}"
            res = solution[key].value
            if isinstance(res, Exception): return 0.0
            try: return float(res)
            except:
                try: return float(res.item())
                except: return 0.0
        
        def safe_float(v):
            try: return float(v)
            except: return 0.0

        # --- RESULTADOS PRINCIPALES ---
        st.divider()
        st.subheader("📊 Resultados de Seguridad")
        
        col_res1, col_res2, col_res3 = st.columns(3)
        k104, l104 = safe_float(get("K104")), safe_float(get("L104"))
        k105, l105 = safe_float(get("K105")), safe_float(get("L105"))
        k106, l106 = safe_float(get("K106")), safe_float(get("L106"))
        k107, l107 = safe_float(get("K107")), safe_float(get("L107"))
        i109, k109 = safe_float(get("I109")), safe_float(get("K109"))
        i110, k110 = safe_float(get("I110")), safe_float(get("K110"))
        
        with col_res1:
            st.markdown("##### ↔️ Desliz. Transversal")
            st.metric("Estribor", f"{k104:.2f} / {l104:.2f}", "OK" if k104 > l104 else "FALLO")
            st.metric("Babor", f"{k105:.2f} / {l105:.2f}", "OK" if k105 > l105 else "FALLO")
        with col_res2:
            st.markdown("##### ↕️ Desliz. Longitudinal")
            st.metric("Proa", f"{k106:.2f} / {l106:.2f}", "OK" if k106 > l106 else "FALLO")
            st.metric("Popa", f"{k107:.2f} / {l107:.2f}", "OK" if k107 > l107 else "FALLO")
        with col_res3:
            st.markdown("##### 🔄 Vuelco")
            st.metric("Estribor", f"{i109:.2f} / {k109:.2f}", "OK" if i109 > k109 else "FALLO")
            st.metric("Babor", f"{i110:.2f} / {k110:.2f}", "OK" if i110 > k110 else "FALLO")

        # --- PANEL DE CONTROL ---
        st.divider()
        st.header("🔍 4. Panel de Control (Valores Internos)")
        
        c_ctrl1, c_ctrl2 = st.columns(2)
        with c_ctrl1:
            st.markdown("##### Fuerzas Estribor (Er)")
            st.write(f"CS Er (D86): {get('D86'):.2f}")
            st.write(f"CS*fy Er (K92): {get('K92'):.2f}")
            st.write(f"CS*c Er (N92): {get('N92'):.2f}")
            
        with c_ctrl2:
            st.markdown("##### Fuerzas Babor (Br)")
            st.write(f"CS Br (D93): {get('D93'):.2f}")
            st.write(f"CS*fy Br (K99): {get('K99'):.2f}")
            st.write(f"CS*c Br (N99): {get('N99'):.2f}")

        st.markdown("##### Fuerzas Long y Vuelco")
        st.write(f"CS*fx Pr (D100): {get('D100'):.2f}")
        st.write(f"CS*fx Pp (G100): {get('G100'):.2f}")
        
        # --- INSPECTOR DE FÓRMULAS EXCEL (NUEVO) ---
        st.divider()
        st.subheader("🕵️ Inspector de Fórmulas Excel (Diagnóstico)")
        st.info("Aquí puedes ver la fórmula real escrita dentro del archivo Excel para detectar diferencias.")
        
        try:
            # Leemos el archivo real para ver las fórmulas (data_only=False)
            wb = openpyxl.load_workbook(ARCHIVO_EXCEL, data_only=False)
            ws = wb[NOMBRE_HOJA]
            
            f_d86 = ws['D86'].value
            f_d93 = ws['D93'].value
            
            col_f1, col_f2 = st.columns(2)
            col_f1.error(f"Fórmula en D86 (Estribor): {f_d86}")
            col_f2.error(f"Fórmula en D93 (Babor): {f_d93}")
            
            if f_d86 != f_d93:
                st.warning("⚠️ ¡ATENCIÓN! Las fórmulas son diferentes. Debes corregir el Excel en la celda D93.")
            else:
                st.success("Las fórmulas parecen idénticas. Revisa referencias circulares o celdas ocultas.")
                
        except Exception as e:
            st.warning(f"No se pudieron leer las fórmulas: {e}")

    except Exception as e:
        st.error(f"Error detallado: {e}")
    
